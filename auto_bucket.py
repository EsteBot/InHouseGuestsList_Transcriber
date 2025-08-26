import streamlit as st
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment, Font
from datetime import date
from io import BytesIO

st.set_page_config(layout="wide", initial_sidebar_state="expanded")

# Create three columns
left_col, middle_col, right_col = st.columns([1, 3, 1])  # [1,2,1] makes middle column wider

# All content will be in the middle column
with middle_col:
    # --------- HEADER UI STYLING ---------
    st.markdown("""
    <style>
    .center { display: flex; justify-content: center; text-align: center; }
    </style>
    """, unsafe_allow_html=True)

    st.markdown("<h2 class='center' style='color:rgb(70, 130, 255);'>An EsteStyle Streamlit Page<br>Where Python Wiz Meets Data Biz!</h2>", unsafe_allow_html=True)
    st.markdown("<img src='https://1drv.ms/i/s!ArWyPNkF5S-foZspwsary83MhqEWiA?embed=1&width=307&height=307' width='300' style='display: block; margin: 0 auto;'>", unsafe_allow_html=True)
    st.markdown("<h3 class='center' style='color: rgb(135, 206, 250);'>🏨 Originally created for Best Western at Firestone 🛎️</h3>", unsafe_allow_html=True)
    st.markdown("<h3 class='center' style='color: rgb(135, 206, 250);'>🤖 By Esteban C Loetz 📟</h3>", unsafe_allow_html=True)
    st.markdown("##")
    st.markdown("---")
    st.markdown("<h3 class='center' style='color: rgb(112, 128, 140);'>📄 In-House Guest List Transcriber 🖋️</h3>", unsafe_allow_html=True)
    st.write("")
    st.markdown("""
    ### 📋 Steps to Export Guest List:
    1. Open the '**Front Office**' user tab
    2. Select '**Reports**' from the top navigation bar
    3. Click '**Front Office**' tab that appears
    4. Hover over '**Reports**' (Bar Graph Icon) at right of screen
    5. Select '**In House Guest**' from dropdown menu
    6. Click '**Refresh**' button
    7. Click '**Export**' button
    8. Select '**Excel**'
    
    #### 📁 An '**In_House_Guests**' file will be created in your Downloads folder.
    #### 🔄 Use this file as input for the transcription automation below.
    """)
    st.markdown("---")

def apply_excel_formatting(ws, guest_data_dict):
    """Formats the new workbook and populates data."""
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    vertical_only_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style=None), bottom=Side(style=None)
    )
    center_aligned_text = Alignment(horizontal='center', vertical='center')

    # Set column widths
    col_widths = {
        'A': 7, 'B': 31, 'C': 10, 'D': 5, 'E': 5,'F': 3, 
        'G': 7, 'H': 31, 'I': 10, 'J': 5, 'K': 5
    }
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    # Set a consistent height for data rows (Rows 2 to 32)
    for row_num in range(2, 34):
        ws.row_dimensions[row_num].height = 14.5

    # Loop through the cells once to apply border and alignment
    # Skip row 2 (the empty row) when applying borders
    for row in range(3, 34):  # Start from row 3 to skip the empty row
        for col in range(1, 12):
            cell = ws.cell(row=row, column=col)
            # Apply special border for column 6 (the empty separator column)
            if col == 6:
                cell.border = vertical_only_border
            else:
                cell.border = thin_border
            cell.alignment = center_aligned_text

    # Add Date and Headers
    bold_font = Font(bold=True)
    
    # Apply date with bold formatting
    ws['B1'] = 'GUEST LIST DATE:'
    ws['B1'].alignment = Alignment(horizontal='right')
    
    ws['C1'] = date.today().strftime('%Y-%m-%d')
    ws['C1'].font = bold_font
    
    headers = ['ROOM', 'GUEST NAME', 'RATE', 'PET', 'INIT', '',
               'ROOM', 'GUEST NAME', 'RATE', 'PET', 'INIT']
    start_col, start_row = 1, 3  # Changed from 2 to 3 to add empty line
    for i, header in enumerate(headers):
        cell = ws.cell(row=start_row, column=start_col + i, value=header)
        cell.font = bold_font

    # Room numbers for column A (Left side)
    room_numbers_col_1 = [105, 106, 107, 108, 109, 110, 111, 112, 114, 115,
                          201, 202, 203, 204, 205, 206, 207, 208, 209, 210,
                          211, 212, 214, 215, 216, 217, 218, 219, 220, 221]
    # Room numbers for column H (Right side)
    room_numbers_col_8 = [222, 223, 224, 225, 226,
                          301, 302, 303, 304, 305, 306, 307, 308, 309, 310,
                          311, 312, 314, 315, 316, 317, 318, 319, 320,
                          321, 322, 323, 324, 325, 326]

    start_row_rooms = 4

    # Populate room numbers for both columns
    for i, room_number in enumerate(room_numbers_col_1):
        cell = ws.cell(row=start_row_rooms + i, column=1, value=room_number)
        cell.font = bold_font
    for i, room_number in enumerate(room_numbers_col_8):
        cell = ws.cell(row=start_row_rooms + i, column=7, value=room_number)
        cell.font = bold_font

    # Populate guest data (Part 4)
    for row_idx in range(3, 33):
        # Column A (Room numbers, starting in row 3)
        room_to_match_col_1 = ws.cell(row=row_idx, column=1).value
        if room_to_match_col_1 in guest_data_dict:
            guest_info = guest_data_dict[room_to_match_col_1]
            ws.cell(row=row_idx, column=2, value=guest_info['Guest_Name']) # Col B (Name)
            ws.cell(row=row_idx, column=3, value=guest_info['Rate'])       # Col C (Rate)

        # Column H (Room numbers, starting in row 3)
        room_to_match_col_7 = ws.cell(row=row_idx, column=7).value
        if room_to_match_col_7 in guest_data_dict:
            guest_info = guest_data_dict[room_to_match_col_7]
            ws.cell(row=row_idx, column=8, value=guest_info['Guest_Name']) # Col H (Name)
            ws.cell(row=row_idx, column=9, value=guest_info['Rate'])      # Col I (Rate)

    return ws

# --- MAIN APP LOGIC ---
with middle_col:
    st.subheader("📥 Download file for transcription:")
    st.write('')
    st.markdown("Upload your **'In_House_Guests.xls'** file to generate the final, formatted guest list as a new `.xlsx` file.")
    st.markdown("This eliminates the need for manual pen & paper transcription!")
    # Part 1: STREAMLIT FILE UPLOAD
    uploaded_file = st.file_uploader("",type=['xls', 'xlsx'] # Streamlit and Pandas can handle both
    )

    if uploaded_file is not None:
        try:
            # Part 2: READ DATA USING PANDAS (Cleaner alternative to xlrd)
            # We specify the header and start row based on your original logic (start at row 16, index 15)
            df = pd.read_excel(
                uploaded_file,
                sheet_name='Sheet1',
                header=None,  # No header row in source data
                skiprows=15   # Skip the first 15 rows (to start at row 16 / index 15)
            )
            st.success(f"Successfully read data from **{uploaded_file.name}**.")

            # Filter the DataFrame using the 'Total Rooms' stop condition
            stop_row_index = df[df.iloc[:, 3].astype(str).str.contains('Total Rooms', na=False)].index
            if not stop_row_index.empty:
                df = df.iloc[:stop_row_index[0]]

            # Keep only the essential columns (D, G, P in original Excel -> index 3, 6, 15 in Python)
            # Note: Since we skipped 15 rows, the columns are now 0, 3, and 12 in the new, smaller DataFrame 'df'
            # Original Excel Cols: [Room (D), Guest Name (G), Rate (P)]
            # New DF Index: [3, 6, 15] - 3 = [0, 3, 12] in the skipped-row DataFrame

            df = df.iloc[:, [3, 6, 15]].copy()
            df.columns = ['Room_Raw', 'Guest_Name', 'Rate_Raw']

            # Clean and process the columns
            # 1. Fill any NaN/blank values in 'Room_Raw' with a placeholder string ('0-')
            df['Room_Raw'] = df['Room_Raw'].fillna('0-').astype(str)

            # 2. Extract the room number (before the dash)
            df['Room_Number'] = df['Room_Raw'].str.split('-').str[0]

            # 3. Convert to integer (this is now safe because we filled the blanks with '0')
            df['Room_Number'] = df['Room_Number'].astype(int)

            # 4. Remove any placeholder rows added in step 1 (where Room_Number is 0)
            df.drop(df[df['Room_Number'] == 0].index, inplace=True)
            # Clean the Rate column: remove '$' and convert to float (handling NaNs/empties)
            df['Rate'] = df['Rate_Raw'].astype(str).str.strip('$').replace('', '0.0').astype(float)
            df.dropna(subset=['Room_Number'], inplace=True) # Remove any rows where Room_Number is missing

            # Build the dictionary (Part 2)
            guest_data_dict = df.set_index('Room_Number')[['Guest_Name', 'Rate']].to_dict('index')
            st.info(f"Loaded **{len(guest_data_dict)}** guest records.")

            # Part 3 & 4: CREATE AND POPULATE NEW WORKBOOK
            new_wb = Workbook()
            new_ws = new_wb.active
            new_ws.page_setup.orientation = new_ws.ORIENTATION_LANDSCAPE
            new_ws = apply_excel_formatting(new_ws, guest_data_dict)

            # Part 5: PREPARE AND SAVE THE FINAL WORKBOOK (for download)
            output = BytesIO()
            new_wb.save(output)
            processed_data = output.getvalue()
            final_file_name = f"In House Guest List {date.today().strftime('%Y%m%d')}.xlsx"

            if st.download_button(
                label="Download Final Guest List (.xlsx)",
                data=processed_data,
                file_name=final_file_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                help="Click to download the newly created and formatted guest list file."
            ):
                st.snow()  # Let it snow when the file is downloaded!
                st.write("<span style='color: rgb(0, 191, 255);'>That's 🥶 COOL baby 🍼! </span>", unsafe_allow_html=True)
            else:
                st.balloons()  # Show balloons when the file is first loaded

        except KeyError as e:
            st.error(f"Error: Could not find the expected data. Please check the **'Sheet1'** sheet or the structure of the file: {e}")
        except ValueError as e:
            st.error(f"Error: Data conversion failed. Check if Room Numbers are valid integers or Rates are valid numbers. Details: {e}")
        except Exception as e:
            st.error(f"An unexpected error occurred during processing: {e}")

    else:
        st.info("Awaiting file upload... Once uploaded, your new file will be ready for download below.")
